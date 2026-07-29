from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def parse_demo(workbook_path: Path) -> dict[str, Any]:
    workbook_path = workbook_path.expanduser().resolve()
    workbook = load_workbook(workbook_path, data_only=False)
    sheets: list[dict[str, Any]] = []

    for worksheet in workbook.worksheets:
        cells: list[dict[str, Any]] = []
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is None or str(cell.value).strip() == "":
                    continue
                cells.append(
                    {
                        "coordinate": cell.coordinate,
                        "row": cell.row,
                        "column": cell.column,
                        "value": str(cell.value),
                        "number_format": cell.number_format,
                        "style_id": cell.style_id,
                    }
                )

        merged = [str(cell_range) for cell_range in worksheet.merged_cells.ranges]
        row_heights = {
            str(index): dimension.height
            for index, dimension in worksheet.row_dimensions.items()
            if dimension.height is not None
        }
        column_widths = {
            key: dimension.width
            for key, dimension in worksheet.column_dimensions.items()
            if dimension.width is not None
        }
        sheets.append(
            {
                "title": worksheet.title,
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
                "cells": cells,
                "merged_ranges": merged,
                "row_heights": row_heights,
                "column_widths": column_widths,
                "note": "坐标用于理解信息层级和相对关系，不自动视为最终设计坐标。",
            }
        )

    return {
        "source_file": workbook_path.name,
        "source_path": str(workbook_path),
        "sheets": sheets,
    }


def write_demo_json(workbook_path: Path, output_path: Path) -> dict[str, Any]:
    result = parse_demo(workbook_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return result
